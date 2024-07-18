from typing import List, Any
import openpyxl

class HelperXlsx:

    def read(self, filename: str) -> List[Any]:
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.active
        data = []

        for row in sheet.iter_rows(values_only=True):
            data.append(list(row))

        return data

    def write(self, filename: str, data: List[Any]):
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        for row in data:
            sheet.append(row)

        workbook.save(filename)

# Nomes posição das colunas
CHANNEL = 0
CHANNEL_CONTACT = 1
AUTOPILOT = 2
SENTIMENT = 3
STARTED = 4
ENDED = 5
DURATION = 6
DURATION_SECONDS = 7
END_OF_INTERACTION = 8
NUMBER_OF_ERRORS = 9
CONVERSATION_MISS = 10
ESCALATION_TIME = 11
INTERACTION_ID = 12
OPCAO_SELECIONADA = 13
JORNADA_DO_CLIENTE_NO_CHATBOT = 14
FINALIZACAO_DO_CONTATO = 15
ANALYZER_ID = 16